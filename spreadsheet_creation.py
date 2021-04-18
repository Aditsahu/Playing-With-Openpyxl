from openpyxl import Workbook

workbook = Workbook()

# Create Worksheet by using Workbook.active property
worksheet = workbook.active

# Create Worksheet by Workbook.create_sheet() method
worksheet1 = workbook.create_sheet("My Sheet")
worksheet2 = workbook.create_sheet("My Sheet", 0)
worksheet3 = workbook.create_sheet("My Sheet", -1)

# Changing the title of worksheet2 ---> 'Worksheet2'
worksheet2.title = "Worksheet2"

# Changing the background colour of title tab
worksheet2.sheet_properties.tabColor = "1072BA"

print(workbook.sheetnames)


# Looping throuugh worksheets
print('Printing Worksheet using loop :')
for sheet in workbook:
    print(sheet)